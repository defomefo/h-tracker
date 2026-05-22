"""
Clean + dedupe the Career Day 2025 sponsors workbook.

Reads:  /Users/defo/Downloads/Partecipanti_Sponsorship 13-14 Nov 2025.xlsx
Writes: scripts/output/sponsors_canonical_2025.csv  — clean, deduped, ingest-ready
        scripts/output/sponsors_review_2025.csv     — flagged rows needing eyes

Strategy:
  1) Sheet 2 (ospiti 1311) is the master company list (one row per company).
  2) Sheet 3 (Adesioni Sponsorship) is the financial + contract sheet.
  3) Sheet 1 (Partecipanti 1411) is structurally garbage (denormalized per-row
     check-in print format) — skipped entirely. Anything missing in 2 + 3 we
     don't need from 1.
  4) Normalize company names: lowercase, strip whitespace, remove punctuation,
     drop trailing legal suffixes (S.p.A., SRL, GmbH, SPA). Use rapidfuzz
     (token_set_ratio) at >= 85 to catch variants ("Bauli" ↔ "Bauli Group",
     "Generali Italia" ↔ "Generali Italia S.p.A", "Diadora" ↔ "Diadora SpA").
  5) Flag H-FARM internal entries (Staff HFC, Studenti HFC, Ospite extra,
     H-Farm AI) so they never leak into the partner database.
  6) Output one row per UNIQUE company merging master + sponsor data, plus a
     separate review CSV for fuzzy-match pairs + missing-data warnings.
"""

import csv
import os
import re
import sys
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook
from rapidfuzz import fuzz, process

# ---------------------------------------------------------------------------
SOURCE_XLSX = "/Users/defo/Downloads/Partecipanti_Sponsorship 13-14 Nov 2025.xlsx"
HERE        = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(HERE, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

CANONICAL_CSV = os.path.join(OUTPUT_DIR, "sponsors_canonical_2025.csv")
REVIEW_CSV    = os.path.join(OUTPUT_DIR, "sponsors_review_2025.csv")

FUZZY_THRESHOLD = 85   # 0-100; rapidfuzz token_set_ratio
EVENT_LABEL     = "Career Day 13-14 Nov 2025"

# ---------------------------------------------------------------------------
# Internal-entry blacklist patterns — these are H-FARM staff/students/booth
# fillers, not partner companies. Substring match, case-insensitive.
INTERNAL_PATTERNS = [
    "staff hfc",
    "studenti hfc",
    "ospite extra",
    "h-farm ai",
    "hfc -",         # generic student team prefix
]

# Common Italian / German legal suffixes to drop during normalization so
# "Bauli S.p.A." ≈ "Bauli". Kept conservative: "Group" is NOT dropped
# (Carraro Group ≠ Carraro a priori; we let fuzzy match handle it).
LEGAL_SUFFIXES = [
    r"\bs\.p\.a\.?",
    r"\bspa\b",
    r"\bs\.r\.l\.?",
    r"\bsrl\b",
    r"\bscpa\b",
    r"\bs\.c\.p\.a\.?",
    r"\bgmbh\b",
    r"\bsb\b",            # società benefit
]

# ---------------------------------------------------------------------------
def normalize_name(raw: str) -> str:
    if not raw:
        return ""
    s = str(raw).strip().lower()
    s = re.sub(r"[ \s]+", " ", s)             # collapse whitespace
    for pat in LEGAL_SUFFIXES:
        s = re.sub(pat, "", s)
    s = re.sub(r"[.,'\"`]", "", s)                 # strip light punctuation
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_internal(raw: str) -> bool:
    if not raw:
        return False
    lo = raw.lower()
    return any(p in lo for p in INTERNAL_PATTERNS)


def is_empty_row(row):
    return all(v in (None, "") for v in row)


# ---------------------------------------------------------------------------
# Sheet 2 — master ospiti
# ---------------------------------------------------------------------------
def parse_master_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    out = OrderedDict()           # norm_name → record (preserves order seen)
    skipped_internal = []
    for r in rows[1:]:
        if is_empty_row(r):
            continue
        raw = (r[1] or "").strip()
        if not raw:
            continue
        if is_internal(raw):
            skipped_internal.append(raw)
            continue
        norm = normalize_name(raw)
        if not norm:
            continue
        # Build attendee list — Ospite 1..5 are cols 6..10 (1-indexed)
        attendees = [v for v in r[5:11] if v]
        rec = {
            "raw_master_name":  raw,
            "area_tematica":    (r[0] or "").strip() if r[0] else "",
            "primary_contact":  (r[2] or "").strip() if r[2] else "",
            "primary_email":    (r[3] or "").strip() if r[3] else "",
            "ospiti_count":     int(r[4]) if r[4] and isinstance(r[4], (int, float)) else None,
            "ospiti_names":     " | ".join(str(a).strip() for a in attendees),
        }
        # If duplicate (same normalized name appears twice in master), keep
        # the entry with more filled-in fields.
        if norm in out:
            existing = out[norm]
            filled_new = sum(1 for v in rec.values() if v)
            filled_old = sum(1 for v in existing.values() if v)
            if filled_new > filled_old:
                out[norm] = rec
        else:
            out[norm] = rec
    return out, skipped_internal


# ---------------------------------------------------------------------------
# Sheet 3 — sponsor
# ---------------------------------------------------------------------------
def parse_sponsor_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    out = OrderedDict()
    skipped_internal = []
    for r in rows[1:]:
        if is_empty_row(r):
            continue
        raw = (r[0] or "").strip()
        if not raw:
            continue
        if is_internal(raw):
            skipped_internal.append(raw)
            continue
        norm = normalize_name(raw)
        if not norm:
            continue
        tier = (r[4] or "").strip() if r[4] else ""
        # tier microcopy normalized → "Base" / "Bronze" / "Gold"
        tier_short = ""
        if "gold" in tier.lower():   tier_short = "Gold"
        elif "bronze" in tier.lower(): tier_short = "Bronze"
        elif "base" in tier.lower():   tier_short = "Base"
        # Partecipazione: cell may be a date (just 14 Nov) or text "entrambe le giornate"
        partecipazione_raw = r[6]
        if hasattr(partecipazione_raw, "strftime"):
            partecipazione = partecipazione_raw.strftime("%Y-%m-%d")
        else:
            partecipazione = str(partecipazione_raw or "").strip()
        rec = {
            "raw_sponsor_name": raw,
            "sponsorship_tier": tier_short,
            "sponsorship_tier_raw": tier,
            "ref_email":        (r[5] or "").strip() if r[5] else "",
            "partecipazione":   partecipazione,
            "note":             (r[7] or "").strip() if r[7] else "",
            "data_pagamento":   r[8].strftime("%Y-%m-%d") if hasattr(r[8], "strftime") else (str(r[8]) if r[8] else ""),
            "value_no_iva":     r[9] if isinstance(r[9], (int, float)) else None,
            "value_with_iva":   r[10] if isinstance(r[10], (int, float)) else None,
            "incassato":        r[11] if isinstance(r[11], (int, float)) else None,
            "signed_by_us":     bool(r[12]) if r[12] is not None else False,
            "signed_by_them":   bool(r[13]) if r[13] is not None else False,
            "fattura_no":       (r[14] or "").strip() if r[14] else "",
            "fattura_date":     r[15].strftime("%Y-%m-%d") if hasattr(r[15], "strftime") else (str(r[15]) if r[15] else ""),
        }
        if norm in out:
            existing = out[norm]
            filled_new = sum(1 for v in rec.values() if v)
            filled_old = sum(1 for v in existing.values() if v)
            if filled_new > filled_old:
                out[norm] = rec
        else:
            out[norm] = rec
    return out, skipped_internal


# ---------------------------------------------------------------------------
def fuzzy_link(sponsor_only_keys, master_keys):
    """For each sponsor-only key, find best fuzzy match in master_keys.
    Returns dict {sponsor_key: (master_key, score)} for matches >= threshold.
    """
    out = {}
    master_list = list(master_keys)
    for sk in sponsor_only_keys:
        match = process.extractOne(sk, master_list, scorer=fuzz.token_set_ratio)
        if match and match[1] >= FUZZY_THRESHOLD:
            out[sk] = (match[0], match[1])
    return out


# ---------------------------------------------------------------------------
def main():
    wb = load_workbook(SOURCE_XLSX, data_only=True, read_only=False)
    print(f"Loaded: {SOURCE_XLSX}")
    print(f"Sheets: {wb.sheetnames}")
    print()

    master, master_skipped = parse_master_sheet(wb["ospiti 1311"])
    sponsor, sponsor_skipped = parse_sponsor_sheet(wb["Adesioni Sponsorship "])
    print(f"Master (ospiti):   {len(master)} unique companies  ·  skipped internal: {len(master_skipped)}")
    print(f"Sponsor (adesioni):{len(sponsor)} unique companies  ·  skipped internal: {len(sponsor_skipped)}")

    if master_skipped:
        print(f"  Master internal entries dropped: {master_skipped}")
    if sponsor_skipped:
        print(f"  Sponsor internal entries dropped: {sponsor_skipped}")

    # ----- Cross-link sponsor → master via exact + fuzzy match -----
    master_keys  = set(master.keys())
    sponsor_keys = set(sponsor.keys())
    exact        = master_keys & sponsor_keys
    sponsor_only = sponsor_keys - master_keys
    master_only  = master_keys - sponsor_keys

    fuzzy_pairs = fuzzy_link(sponsor_only, master_keys)
    print(f"\nMatch breakdown:")
    print(f"  exact (sponsor ∩ master):  {len(exact)}")
    print(f"  fuzzy-linked (≥ {FUZZY_THRESHOLD}): {len(fuzzy_pairs)}")
    print(f"  sponsor-only (no master match):  {len(sponsor_only) - len(fuzzy_pairs)}")
    print(f"  master-only (no sponsor record): {len(master_only)}")

    # ----- Build canonical records keyed by master key when matched, otherwise sponsor key -----
    canonical = OrderedDict()
    review_rows = []

    def merge_record(canon_key, mrec, srec, match_kind=""):
        rec = OrderedDict()
        # Choose display name: prefer master's raw, then sponsor's raw
        display_name = (mrec["raw_master_name"] if mrec else srec["raw_sponsor_name"]) if (mrec or srec) else canon_key
        rec["display_name"]      = display_name
        rec["normalized_name"]   = canon_key
        rec["raw_master_name"]   = mrec["raw_master_name"] if mrec else ""
        rec["raw_sponsor_name"]  = srec["raw_sponsor_name"] if srec else ""
        rec["match_kind"]        = match_kind
        rec["area_tematica"]     = mrec["area_tematica"] if mrec else ""
        rec["primary_contact"]   = mrec["primary_contact"] if mrec else ""
        # email: master > sponsor ref
        rec["primary_email"]     = (mrec["primary_email"] if mrec else "") or (srec["ref_email"] if srec else "")
        rec["ospiti_count"]      = mrec["ospiti_count"] if mrec else None
        rec["ospiti_names"]      = mrec["ospiti_names"] if mrec else ""
        rec["sponsorship_tier"]  = srec["sponsorship_tier"] if srec else ""
        rec["partecipazione"]    = srec["partecipazione"] if srec else ""
        rec["value_no_iva_eur"]  = srec["value_no_iva"] if srec else None
        rec["incassato_eur"]     = srec["incassato"] if srec else None
        rec["signed_by_us"]      = srec["signed_by_us"] if srec else False
        rec["signed_by_them"]    = srec["signed_by_them"] if srec else False
        rec["fattura_no"]        = srec["fattura_no"] if srec else ""
        rec["fattura_date"]      = srec["fattura_date"] if srec else ""
        rec["note"]              = srec["note"] if srec else ""
        rec["event"]             = EVENT_LABEL
        return rec

    # 1) Exact matches → merge master + sponsor
    for k in sorted(exact):
        canonical[k] = merge_record(k, master[k], sponsor[k], match_kind="exact")

    # 2) Fuzzy-linked sponsor-only → merge under master key
    for sk, (mk, score) in fuzzy_pairs.items():
        if mk in canonical:
            # Master already merged via exact; just attach sponsor data
            # (rare — would mean two sponsor rows match one master)
            existing = canonical[mk]
            review_rows.append({
                "issue":   "sponsor-fuzzy-collision",
                "detail":  f"sponsor '{sponsor[sk]['raw_sponsor_name']}' fuzzy-matched master '{master[mk]['raw_master_name']}' (score {score}) but master already merged with another sponsor entry — investigate manually",
                "sponsor_raw": sponsor[sk]['raw_sponsor_name'],
                "master_raw":  master[mk]['raw_master_name'],
            })
            continue
        canonical[mk] = merge_record(mk, master[mk], sponsor[sk], match_kind=f"fuzzy:{score}")
        review_rows.append({
            "issue":   "fuzzy-link-confirm",
            "detail":  f"linked '{sponsor[sk]['raw_sponsor_name']}' → master '{master[mk]['raw_master_name']}' at score {score} — confirm same entity",
            "sponsor_raw": sponsor[sk]['raw_sponsor_name'],
            "master_raw":  master[mk]['raw_master_name'],
        })

    # 3) Sponsor-only with no fuzzy match → sponsor-only canonical
    unmatched_sponsor = sponsor_only - set(fuzzy_pairs.keys())
    for sk in sorted(unmatched_sponsor):
        canonical[sk] = merge_record(sk, None, sponsor[sk], match_kind="sponsor-only")
        review_rows.append({
            "issue":   "sponsor-only-no-match",
            "detail":  f"sponsor '{sponsor[sk]['raw_sponsor_name']}' has no master record — either was sponsor-only or master uses very different name",
            "sponsor_raw": sponsor[sk]['raw_sponsor_name'],
            "master_raw":  "",
        })

    # 4) Master-only → canonical with empty sponsor side (probably visitor not sponsor)
    # Exclude master keys that already got merged via fuzzy in step 2, otherwise
    # we'd overwrite the freshly-merged sponsor data with an empty-sponsor record.
    fuzzy_master_keys = {mk for (mk, _score) in fuzzy_pairs.values()}
    master_only_real  = master_only - fuzzy_master_keys
    for mk in sorted(master_only_real):
        canonical[mk] = merge_record(mk, master[mk], None, match_kind="master-only")
        review_rows.append({
            "issue":   "master-only-no-sponsor",
            "detail":  f"master '{master[mk]['raw_master_name']}' has no sponsor record — visitor only, did not pay?",
            "sponsor_raw": "",
            "master_raw":  master[mk]['raw_master_name'],
        })

    # ----- Additional data-quality flags on the canonical set -----
    for k, rec in canonical.items():
        if not rec["primary_email"]:
            review_rows.append({
                "issue":  "missing-email",
                "detail": f"no email for '{rec['display_name']}'",
                "sponsor_raw": rec["raw_sponsor_name"],
                "master_raw":  rec["raw_master_name"],
            })
        if rec["sponsorship_tier"] and not rec["area_tematica"]:
            review_rows.append({
                "issue":  "missing-area-tematica",
                "detail": f"sponsor '{rec['display_name']}' has no industry sector classified",
                "sponsor_raw": rec["raw_sponsor_name"],
                "master_raw":  rec["raw_master_name"],
            })
        if rec["sponsorship_tier"] and not rec["signed_by_us"] and not rec["signed_by_them"]:
            review_rows.append({
                "issue":  "contract-unsigned-both-sides",
                "detail": f"'{rec['display_name']}' (tier: {rec['sponsorship_tier']}) — neither side signed yet",
                "sponsor_raw": rec["raw_sponsor_name"],
                "master_raw":  rec["raw_master_name"],
            })

    # ----- Write canonical CSV -----
    if canonical:
        fieldnames = list(next(iter(canonical.values())).keys())
        with open(CANONICAL_CSV, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for rec in canonical.values():
                w.writerow(rec)
        print(f"\nWrote canonical CSV: {CANONICAL_CSV}  ({len(canonical)} rows)")

    # ----- Write review CSV -----
    if review_rows:
        with open(REVIEW_CSV, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["issue", "detail", "sponsor_raw", "master_raw"])
            w.writeheader()
            for r in review_rows:
                w.writerow(r)
        print(f"Wrote review CSV:    {REVIEW_CSV}  ({len(review_rows)} flags)")

    # ----- Console summary -----
    by_match = defaultdict(int)
    for rec in canonical.values():
        by_match[rec["match_kind"].split(":")[0]] += 1
    print(f"\n--- Canonical breakdown by match kind ---")
    for kind, count in sorted(by_match.items()):
        print(f"  {kind:<14}  {count}")

    by_tier = defaultdict(int)
    by_tier_value = defaultdict(float)
    for rec in canonical.values():
        if rec["sponsorship_tier"]:
            by_tier[rec["sponsorship_tier"]] += 1
            if rec["value_no_iva_eur"]:
                by_tier_value[rec["sponsorship_tier"]] += rec["value_no_iva_eur"]
    print(f"\n--- Sponsorship tier breakdown ---")
    for tier in ("Gold", "Bronze", "Base"):
        if tier in by_tier:
            print(f"  {tier:<6}  {by_tier[tier]:3d} companies  €{by_tier_value[tier]:>10,.0f}")
    total_value = sum(by_tier_value.values())
    print(f"  {'Total':<6}  {sum(by_tier.values()):3d} companies  €{total_value:>10,.0f}")

    by_issue = defaultdict(int)
    for r in review_rows:
        by_issue[r["issue"]] += 1
    print(f"\n--- Review flags breakdown ---")
    for issue, count in sorted(by_issue.items(), key=lambda x: -x[1]):
        print(f"  {issue:<28}  {count}")

    print(f"\nDONE.")


if __name__ == "__main__":
    main()
